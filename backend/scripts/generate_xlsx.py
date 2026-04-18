#!/usr/bin/env python3
"""XLSX Generator for ConvoiaAI. JSON in via stdin, XLSX written to argv[1]."""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main():
    output_path = sys.argv[1]
    data = json.load(sys.stdin)

    wb = Workbook()
    wb.remove(wb.active)

    purple_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    white_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Calibri', size=11, color='333333')
    header_border = Border(bottom=Side(style='thin', color='DDDDDD'))
    zebra_fill = PatternFill(start_color='F8F8FF', end_color='F8F8FF', fill_type='solid')

    for sheet_data in data.get('sheets', []):
        ws = wb.create_sheet(title=sheet_data.get('name', 'Sheet')[:31])  # Excel sheet name limit

        headers = sheet_data.get('headers', [])
        rows = sheet_data.get('rows', [])
        col_widths = sheet_data.get('columnWidths', [])

        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

        for c, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = white_font
            cell.fill = purple_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = header_border

        for r, row in enumerate(rows, 2):
            for c, value in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = normal_font
                if r % 2 == 0:
                    cell.fill = zebra_fill

        if headers:
            ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(json.dumps({'success': True, 'path': output_path}))


if __name__ == '__main__':
    main()
